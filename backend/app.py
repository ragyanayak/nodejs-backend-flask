from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import csv, io, os, datetime
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "POA_Template.XLSX")

# ─────────────────────────────────────────────────────────────────────────────
# ACTIVITY TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────
ACTIVITY_TEMPLATES = {

    "Disk Extension": {
        "pre": [
            {"task": "Server Health Checks",                      "team": "UNIX TEAM", "duration": datetime.time(0, 15), "comments": ""},
            {"task": "Server Luns and Mount points screen shots", "team": "UNIX TEAM", "duration": datetime.time(0, 15), "comments": ""},
            {"task": "Server Process and Ports Screen shot",      "team": "UNIX TEAM", "duration": datetime.time(0, 15), "comments": ""},
        ],
        "activity": [
            {"task": "Add Disk from infra",    "team": "Infra Team", "duration": datetime.time(0, 15), "comments": ""},
            {"task": "Scan the Disk",          "team": "UNIX TEAM",  "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Create PV",              "team": "UNIX TEAM",  "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Extend VG/LV",           "team": "UNIX TEAM",  "duration": datetime.time(0, 15), "comments": ""},
            {"task": "Resize the Mount point", "team": "UNIX TEAM",  "duration": datetime.time(0, 10), "comments": ""},
        ],
        "post": [
            {"task": "Check the filesystem status", "team": "UNIX TEAM", "duration": datetime.time(0, 15), "comments": ""},
        ],
    },

    "Disk Extension with Downtime": {
        "pre": [
            {"task": "Server Health Checks",                      "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Server Luns and Mount points screen shots", "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Server Process and Ports Screen shot",      "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Suppress the Alerts",                       "team": "CS Mon",  "duration": datetime.time(0, 10), "comments": ""},
        ],
        "activity": [
            {"task": "Stop the APP/DB Services",                       "team": "APP/DB Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Take the snapshot of the server",                "team": "CS VDI",      "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Power off the server",                           "team": "CS Unix",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Add Disk from infra",                            "team": "CS VDI",      "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Scan the Disk",                                  "team": "CS Unix",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Create PV",                                      "team": "CS Unix",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Extend VG/LV",                                   "team": "CS Unix",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Resize the Mount point and power on the server", "team": "CS Unix",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Start the APP/DB Services",                      "team": "APP/DB Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Enable the Alerts on the server", "team": "CS Mon", "duration": datetime.time(0, 15), "comments": ""},
        ],
    },

    "Disk Extension with Downtime - DB services": {
        "pre": [
            {"task": "Server Health Checks",                      "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Server Luns and Mount points screen shots", "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Server Process and Ports Screen shot",      "team": "CS Unix", "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Suppress the Alerts",                       "team": "CS Mon",  "duration": datetime.time(0, 10), "comments": ""},
        ],
        "activity": [
            {"task": "Stop the APP Services",                          "team": "Client",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Stop the DB Services",                           "team": "DB Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Take the snapshot of the server",                "team": "CS VDI",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Power off the server",                           "team": "CS Unix", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Add Disk from infra",                            "team": "CS VDI",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Scan the Disk",                                  "team": "CS Unix", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Create PV",                                      "team": "CS Unix", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Extend VG/LV",                                   "team": "CS Unix", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Resize the Mount point and power on the server", "team": "CS Unix", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Start the DB Services",                          "team": "DB Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Start the APP Services",                         "team": "Client",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Enable the Alerts on the server", "team": "CS Mon", "duration": datetime.time(0, 15), "comments": ""},
        ],
    },

    "IP Swap (VM) with Downtime": {
        "pre": [
            {"task": "Server Health Checks",               "team": "Unix Team",       "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Check Server Luns and Mount points", "team": "Unix Team",       "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Monitoring team to Suppress alerts", "team": "Monitoring Team", "duration": datetime.time(0, 10), "comments": ""},
        ],
        "activity": [
            {"task": "Stop the APP/DB Services",                       "team": "Customer",   "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Take the snapshot of the VM",                    "team": "Infra Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Provide the console",                            "team": "Infra Team", "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "From console Swapped the IPs",                   "team": "Unix Team",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Tag the respective VLANs",                       "team": "Infra Team", "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Validate the gateway ping status within server", "team": "Unix Team",  "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Check the server access within VPN",             "team": "Unix Team",  "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Start the APP/DB Services",                      "team": "Customer",   "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Monitoring team to enable the alerts", "team": "Monitoring Team", "duration": datetime.time(0, 15), "comments": ""},
        ],
    },

    "IP Swap (VM) with Downtime - DB services": {
        "pre": [
            {"task": "Server Health Checks",               "team": "Unix Team",       "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Check Server Luns and Mount points", "team": "Unix Team",       "duration": datetime.time(0, 10), "comments": ""},
            {"task": "Monitoring team to Suppress alerts", "team": "Monitoring Team", "duration": datetime.time(0, 10), "comments": ""},
        ],
        "activity": [
            {"task": "Stop the APP Services",                          "team": "Client",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Stop the DB Services",                           "team": "DB Team",    "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Take the snapshot of the VM",                    "team": "Infra Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Provide the console",                            "team": "Infra Team", "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "From console Swapped the IPs",                   "team": "Unix Team",  "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Tag the respective VLANs",                       "team": "Infra Team", "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Validate the gateway ping status within server", "team": "Unix Team",  "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Check the server access within VPN",             "team": "Unix Team",  "duration": datetime.time(0,  5), "comments": "Down Time Required"},
            {"task": "Start the DB Services",                          "team": "DB Team",    "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Start the APP Services",                         "team": "Client",     "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Monitoring team to enable the alerts", "team": "Monitoring Team", "duration": datetime.time(0, 15), "comments": ""},
        ],
    },

    "Reboot the server": {
        "pre": [
            {"task": "Need to take the server level outputs", "team": "Unix Team", "duration": datetime.time(0, 15), "comments": "No Downtime Required"},
            {"task": "Suppress alerts on the server",         "team": "Mon Team",  "duration": datetime.time(0, 15), "comments": "No Downtime Required"},
            {"task": "Take the snapshot of the server",       "team": "VDI Team",  "duration": datetime.time(0, 30), "comments": "No Downtime Required"},
        ],
        "activity": [
            {"task": "Reboot the server", "team": "Unix Team", "duration": datetime.time(0, 20), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Check the Server Status from Putty", "team": "Unix Team", "duration": datetime.time(0, 30), "comments": "No Downtime Required"},
        ],
    },

    "Reboot the server - DB Services": {
        "pre": [
            {"task": "Need to take the server level outputs", "team": "Unix Team", "duration": datetime.time(0, 15), "comments": "No Downtime Required"},
            {"task": "Suppress alerts on the server",         "team": "Mon Team",  "duration": datetime.time(0, 15), "comments": "No Downtime Required"},
            {"task": "Take the snapshot of the server",       "team": "VDI Team",  "duration": datetime.time(0, 30), "comments": "No Downtime Required"},
        ],
        "activity": [
            {"task": "Stop the DB Services",            "team": "Oracle DB Team", "duration": datetime.time(0, 15), "comments": "Down Time Required"},
            {"task": "Enable the Kdump on the servers", "team": "Unix Team",      "duration": datetime.time(0, 20), "comments": "Down Time Required"},
            {"task": "Reboot the server",               "team": "Unix Team",      "duration": datetime.time(0, 20), "comments": "Down Time Required"},
            {"task": "Start the DB services",           "team": "App/DB Team",    "duration": datetime.time(0, 15), "comments": "Down Time Required"},
        ],
        "post": [
            {"task": "Check the Server Status from Putty", "team": "Unix Team", "duration": datetime.time(0, 30), "comments": "No Downtime Required"},
            {"task": "Monitor the DB Services Status",     "team": "DB Team",   "duration": datetime.time(0, 30), "comments": "No Downtime Required"},
        ],
    },

    "CPU Increase": {
        "pre":      [{"task": "Capture current CPU metrics",  "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [{"task": "Increase vCPU on VM",          "team": "VDI Team",  "duration": datetime.time(0, 15), "comments": "No Downtime Required"}],
        "post":     [{"task": "Verify CPU count post change", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },

    "CPU Increase With Downtime": {
        "pre": [{"task": "Capture current CPU metrics", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [
            {"task": "Shutdown and increase vCPU", "team": "VDI Team", "duration": datetime.time(0, 20), "comments": "Down Time Required"},
            {"task": "Power on the VM",            "team": "VDI Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [{"task": "Verify CPU count post change", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },

    "CPU Increase With Downtime - DB Services": {
        "pre": [{"task": "Capture current CPU metrics", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [
            {"task": "Stop DB Services",           "team": "Oracle DB Team", "duration": datetime.time(0, 15), "comments": "Down Time Required"},
            {"task": "Shutdown and increase vCPU", "team": "VDI Team",       "duration": datetime.time(0, 20), "comments": "Down Time Required"},
            {"task": "Power on the VM",            "team": "VDI Team",       "duration": datetime.time(0, 10), "comments": "Down Time Required"},
            {"task": "Start DB Services",          "team": "App/DB Team",    "duration": datetime.time(0, 15), "comments": "Down Time Required"},
        ],
        "post": [{"task": "Verify CPU and DB status", "team": "Unix Team", "duration": datetime.time(0, 15), "comments": "No Downtime Required"}],
    },

    "RAM Increase": {
        "pre":      [{"task": "Capture current RAM metrics", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [{"task": "Increase RAM on VM",          "team": "VDI Team",  "duration": datetime.time(0, 15), "comments": "No Downtime Required"}],
        "post":     [{"task": "Verify RAM post change",      "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },

    "RAM Increase With Downtime": {
        "pre": [{"task": "Capture current RAM metrics", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [
            {"task": "Shutdown and increase RAM", "team": "VDI Team", "duration": datetime.time(0, 20), "comments": "Down Time Required"},
            {"task": "Power on the VM",           "team": "VDI Team", "duration": datetime.time(0, 10), "comments": "Down Time Required"},
        ],
        "post": [{"task": "Verify RAM post change", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },

    "Shutdown the server": {
        "pre":      [{"task": "Notify stakeholders of shutdown", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [{"task": "Gracefully shutdown the server",  "team": "Unix Team", "duration": datetime.time(0, 15), "comments": "Down Time Required"}],
        "post":     [{"task": "Confirm server is powered off",   "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },

    "Shutdown the server - DB Services": {
        "pre": [{"task": "Notify stakeholders of shutdown", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
        "activity": [
            {"task": "Stop DB Services",               "team": "Oracle DB Team", "duration": datetime.time(0, 15), "comments": "Down Time Required"},
            {"task": "Gracefully shutdown the server", "team": "Unix Team",      "duration": datetime.time(0, 15), "comments": "Down Time Required"},
        ],
        "post": [{"task": "Confirm server is powered off", "team": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Downtime Required"}],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# ROLLBACK PLAN TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────
ROLLBACK_TEMPLATES = {
    "Disk Extension": [
        {"task": "Release the disk from server level", "implementer": "Unix Team",  "duration": datetime.time(0, 10), "comments": "No Down Time", "required_teams": "Unix Team"},
        {"task": "Remove the disk from infra level",   "implementer": "Infra Team", "duration": datetime.time(0, 10), "comments": "No Down Time", "required_teams": "Infra Team"},
    ],
    "Disk Extension with Downtime": [
        {"task": "If there is any boot issue on the server, revert the snapshot.", "implementer": "Unix Team/KVM/FCRT", "duration": datetime.time(0, 40), "comments": "Downtime", "required_teams": "Unix/KVM/FCRT"},
        {"task": "Check all the mount points and services",                        "implementer": "Ctrls-Unix",        "duration": datetime.time(0, 20), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "Disk Extension with Downtime - DB services": [
        {"task": "If there is any boot issue on the server, revert the snapshot.", "implementer": "Unix Team/KVM/FCRT", "duration": datetime.time(0, 40), "comments": "Downtime", "required_teams": "Unix/KVM/FCRT"},
        {"task": "Check all the mount points and services",                        "implementer": "Ctrls-Unix",        "duration": datetime.time(0, 20), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "IP Swap (VM) with Downtime": [
        {"task": "VM - If any issue found disable the NIC and revert the changes", "implementer": "CtrlS Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
        {"task": "VM - If any issue found remove the NIC",                         "implementer": "Infra Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Infra Team"},
        {"task": "VM - If still any issue found revert the snapshot",              "implementer": "Infra Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Infra Team"},
    ],
    "IP Swap (VM) with Downtime - DB services": [
        {"task": "VM - If any issue found disable the NIC and revert the changes", "implementer": "CtrlS Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
        {"task": "VM - If any issue found remove the NIC",                         "implementer": "Infra Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Infra Team"},
        {"task": "VM - If still any issue found revert the snapshot",              "implementer": "Infra Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Infra Team"},
    ],
    "Reboot the server": [
        {"task": "If server does not come up, revert the snapshot", "implementer": "Unix Team/VDI Team", "duration": datetime.time(0, 30), "comments": "Downtime", "required_teams": "Unix/VDI Team"},
        {"task": "Verify all services post rollback",               "implementer": "Unix Team",          "duration": datetime.time(0, 15), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "Reboot the server - DB Services": [
        {"task": "If server does not come up, revert the snapshot",    "implementer": "Unix Team/VDI Team", "duration": datetime.time(0, 30), "comments": "Downtime", "required_teams": "Unix/VDI Team"},
        {"task": "Start the DB Services manually if auto-start fails", "implementer": "Oracle DB Team",     "duration": datetime.time(0, 15), "comments": "Downtime", "required_teams": "Oracle DB Team"},
        {"task": "Verify all services post rollback",                  "implementer": "Unix Team",          "duration": datetime.time(0, 15), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "CPU Increase": [
        {"task": "Revert vCPU count to original value", "implementer": "VDI Team",  "duration": datetime.time(0, 15), "comments": "No Down Time", "required_teams": "VDI Team"},
        {"task": "Verify CPU count post rollback",      "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Down Time", "required_teams": "Unix Team"},
    ],
    "CPU Increase With Downtime": [
        {"task": "Shutdown VM and revert vCPU to original value", "implementer": "VDI Team",  "duration": datetime.time(0, 20), "comments": "Downtime", "required_teams": "VDI Team"},
        {"task": "Power on VM and verify CPU count",              "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "CPU Increase With Downtime - DB Services": [
        {"task": "Shutdown VM and revert vCPU to original value", "implementer": "VDI Team",    "duration": datetime.time(0, 20), "comments": "Downtime", "required_teams": "VDI Team"},
        {"task": "Power on VM and start DB Services",             "implementer": "App/DB Team", "duration": datetime.time(0, 15), "comments": "Downtime", "required_teams": "App/DB Team"},
        {"task": "Verify CPU count and DB status post rollback",  "implementer": "Unix Team",   "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "RAM Increase": [
        {"task": "Revert RAM to original value", "implementer": "VDI Team",  "duration": datetime.time(0, 15), "comments": "No Down Time", "required_teams": "VDI Team"},
        {"task": "Verify RAM post rollback",     "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "No Down Time", "required_teams": "Unix Team"},
    ],
    "RAM Increase With Downtime": [
        {"task": "Shutdown VM and revert RAM to original value", "implementer": "VDI Team",  "duration": datetime.time(0, 20), "comments": "Downtime", "required_teams": "VDI Team"},
        {"task": "Power on VM and verify RAM",                   "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "Shutdown the server": [
        {"task": "Power on the server if rollback required",     "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
        {"task": "Verify server is up and services are running", "implementer": "Unix Team", "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
    "Shutdown the server - DB Services": [
        {"task": "Power on the server",                       "implementer": "Unix Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
        {"task": "Start DB Services post power-on",           "implementer": "Oracle DB Team", "duration": datetime.time(0, 15), "comments": "Downtime", "required_teams": "Oracle DB Team"},
        {"task": "Verify server and DB Services are running", "implementer": "Unix Team",      "duration": datetime.time(0, 10), "comments": "Downtime", "required_teams": "Unix Team"},
    ],
}

# ─────────────────────────────────────────────────────────────────────────────
# ROW CONSTANTS — confirmed from template inspection
#
# Activity Plan sheet layout:
#   Row 2          : Pre-Activity Plan header (merged B2:K2)
#   Row 3          : Pre-Activity col headers
#   Rows 4–8       : Pre data slots (up to 5 rows)
#   Row 9          : Activity-Plan header (merged B9:K9)
#   Row 10         : Activity col headers
#   Rows 11–19     : Activity data slots (up to 9 rows)
#   Row 20         : Post Activity Plan header (merged B20:K20)
#   Row 21         : Post-Activity col headers
#   Rows 22–26     : Post data slots (up to 5 rows)
#
# Rollback Plan sheet layout:
#   Row 2          : Rollback Plan header (merged B2:H2)
#   Row 3          : Col headers
#   Rows 4+        : Data rows
# ─────────────────────────────────────────────────────────────────────────────
ROW_PRE_START  = 4
ROW_ACT_START  = 11
ROW_POST_START = 23
ROW_RB_START   = 4

MAX_PRE      = 5
MAX_ACT      = 9
MAX_POST     = 5
MAX_ROLLBACK = 5


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_dt(dt_str):
    if not dt_str:
        return ""
    return str(dt_str).strip().replace("T", " ")


def _parse_datetime(dt_str):
    if not dt_str:
        return "", ""
    s = str(dt_str).strip()
    sep = "T" if "T" in s else " "
    if sep in s:
        parts = s.split(sep, 1)
        return parts[0], parts[1][:5]
    return s, ""


def _add_duration_to_time(time_str, duration_time):
    if not time_str or not isinstance(duration_time, datetime.time):
        return ""
    try:
        parts = str(time_str).strip().split(":")
        h, m = int(parts[0]), int(parts[1])
        total = h * 60 + m + duration_time.hour * 60 + duration_time.minute
        total %= 1440
        return f"{total // 60}:{total % 60:02d}"
    except Exception:
        return ""


def _is_merged(ws, row, col):
    from openpyxl.cell.cell import MergedCell
    return isinstance(ws.cell(row, col), MergedCell)


def _safe_set(ws, row, col, value):
    if not _is_merged(ws, row, col):
        ws.cell(row, col).value = value


def _fmt_duration(t):
    if isinstance(t, datetime.time):
        return f"{t.hour}:{t.minute:02d}"
    return str(t)


def _write_task(ws, row, idx, task, start_datetime_str):
    """
    Activity Plan column mapping (confirmed from template):
    col2=Sl.No | col3=Task | col4=Owner/Team | col5=Status |
    col6=Comments | col7=Duration(Min) | col8=Start Date |
    col9=Start Time | col10=End Time | col11=Remarks
    """
    start_date, start_time = _parse_datetime(start_datetime_str)
    end_time = _add_duration_to_time(start_time, task["duration"])

    _safe_set(ws, row,  2, idx)
    _safe_set(ws, row,  3, task["task"])
    _safe_set(ws, row,  4, task["team"])
    _safe_set(ws, row,  5, "Open")
    _safe_set(ws, row,  6, task["comments"])
    _safe_set(ws, row,  7, _fmt_duration(task["duration"]))
    _safe_set(ws, row,  8, start_date)
    _safe_set(ws, row,  9, start_time)
    _safe_set(ws, row, 10, end_time)
    _safe_set(ws, row, 11, start_date)


def _clear_task(ws, row):
    for c in range(2, 12):
        _safe_set(ws, row, c, None)


def _write_rollback(ws, row, idx, step, planned_date):
    """
    Rollback Plan column mapping (confirmed from template):
    col2=Sl.No | col3=PlannedDate/Time | col4=Duration |
    col5=Implementer | col6=Activity/Step Details |
    col7=Required Teams | col8=Expected Result
    """
    _safe_set(ws, row, 2, idx)
    _safe_set(ws, row, 3, planned_date)
    _safe_set(ws, row, 4, _fmt_duration(step["duration"]))
    _safe_set(ws, row, 5, step["implementer"])
    _safe_set(ws, row, 6, step["task"])
    _safe_set(ws, row, 7, step["required_teams"])
    _safe_set(ws, row, 8, step["comments"])


def _clear_rollback(ws, row):
    for c in range(2, 9):
        _safe_set(ws, row, c, None)


# ── Sheet fillers ─────────────────────────────────────────────────────────────

def fill_header(ws, d):
    """
    Header sheet confirmed cell mapping:
    C3=Date, E3=Requestor(client_spoc)
    C4=CR_ID, E4=Department(Client)
    C6=Org Name, E6=Customer Email
    C7=Customer Name(client_spoc), E7=Yes
    C9=Change Description, C10=Change Justification
    C11=Start datetime, E11=End datetime
    C12=Duration, D12=Downtime
    B14:C15=Risk CtrlS value, D14:E15=Risk Customer value
    C17=Rollback reviewer, C18=Rollback responsible
    D20=Name of Approver (COE Manager designation already in C20)
    E20=pm_name
    C21=Name of Approver/Reviewer
    C22=Change Performer
    C23=Change Review - Post Activity
    """
    approver = d.get("approver_fixed", "Kishore Kumar Manne")
    if d.get("approver_extra", "").strip():
        approver += f", {d['approver_extra'].strip()}"

    post_rev = d.get("post_review_fixed", "Kishore Kumar Manne")
    if d.get("post_review_extra", "").strip():
        post_rev += f", {d['post_review_extra'].strip()}"

    # Row 3: Date and Requestor
    ws["C3"] = d.get("cr_date", "")
    ws["E3"] = d.get("client_spoc", "")

    # Row 4: CR ID and Department
    ws["C4"] = d.get("cr_id", "")
    ws["E4"] = "Client"

    # Row 6: Organization Name and Customer Email
    ws["C6"] = d.get("client_name", "")
    ws["E6"] = d.get("client_mail", "")

    # Row 7: Customer Name and approval flag
    ws["C7"] = d.get("client_spoc", "")
    ws["E7"] = "Yes"

    # Row 9: Change Description
    ws["C9"] = f"{d.get('activity', '')} - {d.get('client_name', '')}"

    # Row 10: Change Justification
    ws["C10"] = d.get("risk", "")

    # Row 11: Change Start/End datetime
    ws["C11"] = _fmt_dt(d.get("start_datetime", ""))
    ws["E11"] = _fmt_dt(d.get("end_datetime", ""))

    # Row 12: Duration
    ws["C12"] = d.get("duration", "")

    # Rows 14-15: Risk values (merged areas)
    _safe_set(ws, 14, 2, d.get("risk", ""))
    _safe_set(ws, 14, 4, d.get("risk", ""))

    # Row 17: Rollback reviewer (value col = C)
    ws["C17"] = d.get("rollback_reviewer", "")

    # Row 18: Rollback responsible (value col = C)
    ws["C18"] = d.get("rollback_responsible", "")

    # Row 20: COE Manager designation + Name of Approver
    ws["C20"] = "COE Manager - Unix"
    ws["E20"] = d.get("pm_name", "")

    # Row 21: Name of Approver/Reviewer
    ws["C21"] = approver

    # Row 22: Change Performer
    ws["C22"] = d.get("change_performer", "")

    # Row 23: Change Review - Post Activity
    ws["C23"] = post_rev


def fill_activity_plan(ws, d):
    activity = d.get("activity", "Reboot the server")
    tmpl = ACTIVITY_TEMPLATES.get(activity, ACTIVITY_TEMPLATES["Reboot the server"])
    start_dt_str = d.get("start_datetime", "")

    # Pre-Activity rows (start at row 4)
    pre_tasks = tmpl["pre"]
    for i in range(MAX_PRE):
        row = ROW_PRE_START + i
        if i < len(pre_tasks):
            _write_task(ws, row, i + 1, pre_tasks[i], start_dt_str)
        else:
            _clear_task(ws, row)

    # Activity rows (start at row 11)
    act_tasks = tmpl["activity"]
    for i in range(MAX_ACT):
        row = ROW_ACT_START + i
        if i < len(act_tasks):
            _write_task(ws, row, i + 1, act_tasks[i], start_dt_str)
        else:
            _clear_task(ws, row)

    # Post-Activity rows (start at row 22)
    post_tasks = tmpl["post"]
    for i in range(MAX_POST):
        row = ROW_POST_START + i
        if i < len(post_tasks):
            _write_task(ws, row, i + 1, post_tasks[i], start_dt_str)
        else:
            _clear_task(ws, row)


def fill_rollback_plan(ws, d):
    activity     = d.get("activity", "")
    planned_date = d.get("activity_start_date", "TBD")
    steps = ROLLBACK_TEMPLATES.get(activity, [
        {
            "task":           "Revert to previous configuration if any issue found",
            "implementer":    "Unix Team",
            "duration":       datetime.time(0, 15),
            "comments":       "Downtime",
            "required_teams": "Unix Team",
        }
    ])

    for i in range(MAX_ROLLBACK):
        row = ROW_RB_START + i
        if i < len(steps):
            _write_rollback(ws, row, i + 1, steps[i], planned_date)
        else:
            _clear_rollback(ws, row)


def fill_server_details(ws, d):
    servers = d.get("servers", [])
    # Clear up to 15 rows
    for i in range(15):
        r = 5 + i
        ws.cell(r, 3).value = None
        ws.cell(r, 4).value = None
        ws.cell(r, 5).value = None
    for i, srv in enumerate(servers):
        r = 5 + i
        ws.cell(r, 3).value = i + 1
        ws.cell(r, 4).value = srv.get("db_name", "")
        ws.cell(r, 5).value = srv.get("ip", "")


def fill_communication_plan(ws, d):
    """
    Communication Plan confirmed layout:
    Row 4: C4=Start datetime (formula =Header!C11), D4=End datetime (formula =Header!E11),
           E4=Duration (formula =D4-C4), F4=Name (formula =Header!C7), H4=Email (formula =Header!E6)
    We write directly to override the formulas with actual values:
    col3=Start, col4=End, col5=Duration, col6=Name, col7=Phone, col8=Email,
    col9=Client Impact, col10=Teams to notify
    """
    _safe_set(ws, 4, 3,  _fmt_dt(d.get("start_datetime",    "")))
    _safe_set(ws, 4, 4,  _fmt_dt(d.get("end_datetime",       "")))
    _safe_set(ws, 4, 5,  d.get("duration",           ""))
    _safe_set(ws, 4, 6,  d.get("comm_name",          ""))
    _safe_set(ws, 4, 7,  d.get("comm_phone",         ""))
    _safe_set(ws, 4, 8,  d.get("comm_email",         ""))
    _safe_set(ws, 4, 9,  d.get("comm_client_impact", ""))
    _safe_set(ws, 4, 10, d.get("comm_teams_notify",  ""))


def fill_escalation_matrix(ws, d):
    """
    Escalation Matrix confirmed layout:
    Row 2: header (merged B2:F2) = "Ctrls Escalation"
    Row 3: col headers
    Rows 4+: CtrlS escalation data
    Row 10: "Client Escalation" header (merged B10:F10)
    Row 11: col headers
    Row 12: client data (formulas =Header!C6, =Header!C7, =Header!E6)
    Rows 13+: additional escalation rows
    """
    esc_rows = d.get("escalation_rows", [])

    # CtrlS escalation rows start at row 4
    for i, row in enumerate(esc_rows):
        r = 4 + i
        _safe_set(ws, r, 2, row.get("resource",   ""))
        _safe_set(ws, r, 3, row.get("executor",   ""))
        _safe_set(ws, r, 4, row.get("escalation", ""))
        _safe_set(ws, r, 5, row.get("contact",    ""))
        _safe_set(ws, r, 6, row.get("oncall",     ""))

    # Client escalation: row 12 = client info (override formulas)
    _safe_set(ws, 12, 2, d.get("client_name",  ""))
    _safe_set(ws, 12, 3, "Client")
    _safe_set(ws, 12, 4, d.get("client_spoc",  ""))
    _safe_set(ws, 12, 5, d.get("client_mail",  ""))
    _safe_set(ws, 12, 6, "")

    # Additional CtrlS rows mirrored under client escalation
    for i, row in enumerate(esc_rows):
        r = 13 + i
        _safe_set(ws, r, 2, row.get("resource",   ""))
        _safe_set(ws, r, 3, row.get("executor",   ""))
        _safe_set(ws, r, 4, row.get("escalation", ""))
        _safe_set(ws, r, 5, row.get("contact",    ""))
        _safe_set(ws, r, 6, row.get("oncall",     ""))


# ── CSV generation ─────────────────────────────────────────────────────────────

def _generate_csv(d):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["POA DETAILS"])
    w.writerow([])
    for label, key in [
        ("Date",                       "cr_date"),
        ("CR ID",                      "cr_id"),
        ("Client Name",                "client_name"),
        ("Client SPOC",                "client_spoc"),
        ("Client Mail",                "client_mail"),
        ("Client Mobile",              "client_mobile"),
        ("Change Start Datetime",      "start_datetime"),
        ("Change End Datetime",        "end_datetime"),
        ("Duration (Change End - Change Start)", "duration"),
        ("Activity",                   "activity"),
        ("Infra Team",                 "infra_team"),
        ("PM Name",                    "pm_name"),
        ("PM Mail",                    "pm_mail"),
        ("PM Mobile",                  "pm_mobile"),
        ("Risk",                       "risk"),
        ("Rollback Reviewer",          "rollback_reviewer"),
        ("Rollback Responsible",       "rollback_responsible"),
        ("Approver / Reviewer",        "approver_fixed"),
        ("Approver Extra",             "approver_extra"),
        ("Change Performer",           "change_performer"),
        ("Post Activity Review",       "post_review_fixed"),
        ("Post Activity Review Extra", "post_review_extra"),
        ("Comm Name",                  "comm_name"),
        ("Comm Phone",                 "comm_phone"),
        ("Comm Email",                 "comm_email"),
        ("Comm Client Impact",         "comm_client_impact"),
        ("Comm Teams Notify",          "comm_teams_notify"),
    ]:
        val = d.get(key, "")
        if key in ("start_datetime", "end_datetime"):
            val = _fmt_dt(val)
        w.writerow([label, val])

    activity = d.get("activity", "")
    start_dt = d.get("start_datetime", "TBD")
    start_date, start_time = _parse_datetime(start_dt)

    act_tmpl = ACTIVITY_TEMPLATES.get(activity, {"pre": [], "activity": [], "post": []})
    for section, tasks in [
        ("Pre-Activity",  act_tmpl["pre"]),
        ("Activity",      act_tmpl["activity"]),
        ("Post-Activity", act_tmpl["post"]),
    ]:
        w.writerow([])
        w.writerow([section])
        w.writerow(["#", "Task", "Team", "Status", "Comments", "Duration",
                    "Start Date", "Start Time", "End Time"])
        for i, t in enumerate(tasks, 1):
            end_time = _add_duration_to_time(start_time, t["duration"])
            w.writerow([i, t["task"], t["team"], "Open", t["comments"],
                        _fmt_duration(t["duration"]), start_date, start_time, end_time])

    rb_steps = ROLLBACK_TEMPLATES.get(activity, [])
    w.writerow([])
    w.writerow(["ROLLBACK PLAN"])
    w.writerow(["#", "Planned Date/Time", "Duration", "Implementer",
                "Activity/Step Details", "Required Teams", "Comments"])
    for i, s in enumerate(rb_steps, 1):
        w.writerow([i, d.get("activity_start_date", "TBD"), _fmt_duration(s["duration"]),
                    s["implementer"], s["task"], s["required_teams"], s["comments"]])

    w.writerow([])
    w.writerow(["SERVER DETAILS"])
    w.writerow(["#", "Environment", "DB Name", "IP", "Kdump"])
    for i, s in enumerate(d.get("servers", []), 1):
        w.writerow([i, s.get("env", ""), s.get("db_name", ""),
                    s.get("ip", ""), s.get("kdump", "")])

    w.writerow([])
    w.writerow(["CTRLS ESCALATION"])
    w.writerow(["Resource Summary", "Executor", "Escalation", "Contact Details", "On Site On Call"])
    for row in d.get("escalation_rows", []):
        w.writerow([row.get("resource",""), row.get("executor",""), row.get("escalation",""),
                    row.get("contact",""), row.get("oncall","")])

    w.writerow([])
    w.writerow(["CLIENT ESCALATION"])
    w.writerow(["Resource Summary", "Executor", "Escalation", "Contact Details", "On Site On Call"])
    w.writerow([d.get("client_name",""), "Client", d.get("client_spoc",""), d.get("client_mail",""), ""])
    for row in d.get("escalation_rows", []):
        w.writerow([row.get("resource",""), row.get("executor",""), row.get("escalation",""),
                    row.get("contact",""), row.get("oncall","")])

    return buf.getvalue()


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/generate-excel", methods=["POST"])
def generate_excel_route():
    d = request.json

    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({
            "error": f"Template not found: {TEMPLATE_PATH}. "
                     "Place POA_Template.XLSX next to app.py."
        }), 500

    with open(TEMPLATE_PATH, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()))

    fill_header(wb["Header"], d)
    fill_activity_plan(wb["Activity Plan"], d)
    fill_rollback_plan(wb["Rollback Plan"], d)
    fill_server_details(wb["Server Details"], d)
    fill_communication_plan(wb["Communication Plan"], d)
    fill_escalation_matrix(wb["Escalation Matrix"], d)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    cr_id = d.get("cr_id", "output").replace(" ", "_")
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"POA_{cr_id}.xlsx",
    )


@app.route("/generate-csv", methods=["POST"])
def generate_csv_route():
    d = request.json
    content = _generate_csv(d)
    out = io.BytesIO(content.encode("utf-8-sig"))
    out.seek(0)
    cr_id = d.get("cr_id", "output").replace(" ", "_")
    return send_file(out, mimetype="text/csv", as_attachment=True,
                     download_name=f"POA_{cr_id}.csv")


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "template_exists": os.path.exists(TEMPLATE_PATH)})


@app.route("/activities", methods=["GET"])
def get_activities():
    return jsonify(list(ACTIVITY_TEMPLATES.keys()))


# ── Debug endpoints ────────────────────────────────────────────────────────────

@app.route("/debug-header", methods=["GET"])
def debug_header():
    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({"error": "Template not found"}), 500
    with open(TEMPLATE_PATH, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()))
    ws = wb["Header"]
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append({"address": cell.coordinate, "row": cell.row,
                               "col": cell.column, "value": str(cell.value)[:80]})
    return jsonify(cells)


@app.route("/debug-activity", methods=["GET"])
def debug_activity():
    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({"error": "Template not found"}), 500
    with open(TEMPLATE_PATH, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()))
    ws = wb["Activity Plan"]
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append({"address": cell.coordinate, "row": cell.row,
                               "col": cell.column, "value": str(cell.value)[:80]})
    merged = [str(r) for r in ws.merged_cells.ranges]
    return jsonify({"cells": cells, "merged_ranges": merged})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
